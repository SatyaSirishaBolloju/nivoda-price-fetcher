import requests
import openpyxl


# Load Excel workbook and sheet
def load_excel(file_path):
    """Load the Excel workbook and first sheet."""
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    return wb, ws

# Read config from txt file
def load_config(config_path='nivoda_excel.txt'):
    """Load Excel file path, diamond count, and API token from config."""
    config = {}
    with open(config_path, 'r') as f:
        for line in f:
            if '=' in line:
                key, value = line.strip().split('=')
                config[key.strip()] = value.strip()
    return config

# Build GraphQL query for a diamond based on row data
def build_query(shape, weight, color, clarity):
    """Create GraphQL query payload for diamond listing."""
    return {
        "query": """
        query SearchStones($pagination: PaginationInput, $filters: StoneFilterInput) {
          searchStones(pagination: $pagination, filters: $filters) {
            edges {
              node {
                bestPrice
                bestPricePerCarat
                color
                clarity
                shape
                weight
              }
            }
          }
        }
        """,
        "variables": {
            "pagination": {
                "page": 1,
                "limit": 1
            },
            "filters": {
                "color": [color],
                "clarity": [clarity],
                "shape": shape,
                "weight": {
                    "from": float(weight) - 0.02,
                    "to": float(weight) + 0.02
                }
            }
        }
    }

# Send GraphQL request to Nivoda API
def fetch_price(token, query):
    """Send a POST request to Nivoda's GraphQL endpoint and return price info."""
    headers = {
        'Content-Type': 'application/json',
        'Authorization': f'Bearer {token}'  # 🔒 Do not publish real tokens
    }
    try:
        response = requests.post("https://api.nivoda.net/graphql", headers=headers, json=query)
        response.raise_for_status()
        data = response.json()
        results = data['data']['searchStones']['edges']
        if results:
            return results[0]['node']['bestPrice'], results[0]['node']['bestPricePerCarat']
    except Exception as e:
        print(f"Error fetching price: {e}")
    return None, None

# Main logic to loop through rows, fetch price, and write back
def process_excel(file_path, token, diamond_count):
    """Process Excel row-by-row, fetch price from Nivoda, and write results."""
    wb, ws = load_excel(file_path)

    for i in range(2, int(diamond_count) + 2):  # Assuming headers in row 1
        shape = ws[f"B{i}"].value
        weight = ws[f"D{i}"].value
        color = ws[f"E{i}"].value
        clarity = ws[f"F{i}"].value

        print(f"Fetching price for Row {i-1}: {shape}, {weight}, {color}, {clarity}")
        query = build_query(shape, weight, color, clarity)
        total_price, price_per_carat = fetch_price(token, query)

        if price_per_carat:
            ws[f"N{i}"].value = price_per_carat
            ws[f"O{i}"].value = total_price

    wb.save(file_path)
    print("✅ Excel file updated successfully.")

# Run everything
if __name__ == "__main__":
    config = load_config()
    file_path = config['file_path']
    diamond_count = config['diamond']
    token = config['token']  # 🔒 REDACT before uploading to GitHub
    process_excel(file_path, token, diamond_count)
